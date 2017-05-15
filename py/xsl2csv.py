__author__ = 'Jason'

from openpyxl import load_workbook
wb = load_workbook("C:\\Users\\Jason\\Desktop\\alldeaf_Health_08_20_2015.xlsx")

ws = wb['Health']

print (ws['A2'])

